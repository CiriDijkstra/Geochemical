from openpyxl import *
from Sample import *
from Element import *
import numpy as np
import time

time_start = time.time()
wb = load_workbook('samples.xlsx')      #务必将存储原始数据的表格文件命名为“samples.xlsx”
sheet = wb.worksheets[0]
#以下程序段用于读入各变量在表格文件中的列数，务必将变量名存储在第一行
#务必将高斯坐标X、Y坐标的变量名存储为“XX”和“YY”
#务必将各元素含量的变量名存储为该元素的符号，如W、Sn、Au等
#务必将类标号的变量名存储为“class”
XX_col, YY_col, W_col, Sn_col, Pb_col, Zn_col, class_col = 0, 0, 0, 0, 0, 0, 0
for col in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=col).value == 'XX':
        XX_col = col
    if sheet.cell(row=1, column=col).value == 'YY':
        YY_col = col
    if sheet.cell(row=1, column=col).value == 'W':
        W_col = col
    if sheet.cell(row=1, column=col).value == 'Sn':
        Sn_col = col
    if sheet.cell(row=1, column=col).value == 'Pb':
        Pb_col = col
    if sheet.cell(row=1, column=col).value == 'Zn':
        Zn_col = col
    if sheet.cell(row=1, column=col).value == 'class':
        class_col = col

#以下程序段将从原始数据表格中导入各样品的信息，包括样品点的高斯坐标、元素含量和类标号等
sample_list = []
for i in range(2,1862):
    sample_one = Sample(sheet.cell(row=i,column=XX_col).value, sheet.cell(row=i,column=YY_col).value,
                    sheet.cell(row=i,column=W_col).value, sheet.cell(row=i,column=Sn_col).value,
                    sheet.cell(row=i,column=Pb_col).value, sheet.cell(row=i,column=Zn_col).value,
                    sheet.cell(row=i,column=class_col).value)
    for j in range(35,42):
        sample_one.macro_element.append(sheet.cell(row=i,column=j).value)
    sample_list.append(sample_one)

#以下程序段将分别计算各微量元素在各个类别中的分布模型和背景上限，结果将输出到表格文件“分类情况.xlsx”
W = ElementList('W')
Sn = ElementList('Sn')
Pb = ElementList('Pb')
Zn = ElementList('Zn')
temp_list = [W, Sn, Pb, Zn]
wb1 = Workbook()
sheet_ca = wb1.worksheets[0]
sheet_ca.append(['类标号', '样品数量', '元素', '分布模型', '背景上限'])
for i in range(1,6):
    for each in sample_list:
        if each.class_id == i:
            W.append(each.W)
            Sn.append(each.Sn)
            Pb.append(each.Pb)
            Zn.append(each.Zn)
    for element in temp_list:
        sheet_ca.append([i, len(element), element.element_name, element.distribution_pattern(), element.c_a()])
        del element[ : ]
wb1.save('分类情况.xlsx')

#以下程序段将各元素在各分类下的背景上限存储到一个字典列表中，列表由5个字典组成，每个字典代表一个分类
#每个字典中的键为元素名称（字符串），值为该元素在该分类下的背景上限
sheet2 = wb1.worksheets[0]
c_a = [{'W': None, 'Sn': None, 'Pb': None, 'Zn': None}, {'W': None, 'Sn': None, 'Pb': None, 'Zn': None},
       {'W': None, 'Sn': None, 'Pb': None, 'Zn': None}, {'W': None, 'Sn': None, 'Pb': None, 'Zn': None},
       {'W': None, 'Sn': None, 'Pb': None, 'Zn': None}]
for i in range(0,5):
    c_a[i]['W'] = sheet2.cell(row=i*4+2, column=5).value
    c_a[i]['Sn'] = sheet2.cell(row=i*4+3, column=5).value
    c_a[i]['Pb'] = sheet2.cell(row=i*4+4, column=5).value
    c_a[i]['Zn'] = sheet2.cell(row=i*4+5, column=5).value

#以下程序段将计算所有样品各个元素的变差衬度值，相邻点的判定条件是空间上相邻且属于同一类
#若一个点没有相邻点，则令其衬度值为0
for each in sample_list:
    neighborhood = []
    for another in sample_list:
        if another != each and another.class_id == each.class_id and another.is_adjacency(each):
           neighborhood.append(another)
    W, Sn, Pb, Zn = [], [], [], []
    for each_neighbor in neighborhood:
        W.append(each_neighbor.W)
        Sn.append(each_neighbor.Sn)
        Pb.append(each_neighbor.Pb)
        Zn.append(each_neighbor.Zn)
    if len(W) > 0:
        each.W_contrast = (each.W - np.mean(W)) / c_a[each.class_id - 1]['W']
    else:
        each.W_contrast = 0
    if len(Sn) > 0:
        each.Sn_contrast = (each.Sn - np.mean(Sn)) / c_a[each.class_id - 1]['Sn']
    else:
        each.Sn_contrast = 0
    if len(Pb) > 0:
        each.Pb_contrast = (each.Pb - np.mean(Pb)) / c_a[each.class_id - 1]['Pb']
    else:
        each.Pb_contrast = 0
    if len(Zn) > 0:
        each.Zn_contrast = (each.Zn - np.mean(Zn)) / c_a[each.class_id - 1]['Zn']
    else:
        each.Zn_contrast = 0

#以下程序段将各样品的高斯坐标、元素变差衬度值输出到表格文件“变差衬度值.xlsx”中
wb2 = Workbook()
sheet_contrast = wb2.worksheets[0]
sheet_contrast.append(['XX', 'YY', 'W', 'Sn', 'Pb', 'Zn'])
for each_sample in sample_list:
    sheet_contrast.append([each_sample.XX, each_sample.YY, each_sample.W_contrast, each_sample.Sn_contrast,
                           each_sample.Pb_contrast, each_sample.Zn_contrast])
wb2.save('变差衬度值.xlsx')

print('分类情况及各区元素背景上限值输出至文件 “分类情况.xlsx” 中\n变差衬度值计算结果输出至文件 “变差衬度值.xlsx” 中')

time_end = time.time()
print('共用时: ', time_end - time_start, 's', sep='')
